
# coding: utf-8

# In[99]:

import openpyxl
import pandas as pd

#filename = input('Enter workbook filename: ')
filename = 'Marvel v2 ProgressionCombatant.xlsx'
wb = openpyxl.load_workbook(filename)

characterlist = 'characterlist.csv'
df = pd.read_csv(characterlist, sep=',',header=0)


# In[100]:

dfmale = df[df['Rig'] == 'Male']
dffemale = df[df['Rig'] == 'Female']
dfsorted = dfmale.append(dffemale)
dfsorted = dfsorted.sort_values('character_id')
print(dfsorted.info())
print(dfsorted.head(5))


# In[101]:

dfsorted = dfsorted.dropna()
dfsorted.info()


# In[102]:

sheet_count = len(wb.get_sheet_names())
print('First ten sheets: ' + str(wb.get_sheet_names()[:10]))
print('Number of sheets: ' + str(sheet_count))


# In[103]:

#target = wb.copy_worksheet(antManCbt)
#newsheetname = wb.get_sheet_names()[-1]
#wb.get_sheet_by_name(newsheetname).title = 'newSheet'
#newsheet = wb.get_sheet_by_name('newSheet')

antManCbt = wb.get_sheet_by_name('antManCbt')
agent13Cbt = wb.get_sheet_by_name('agent13Cbt')

for i in range(0,len(dfsorted)):
    if dfsorted.values[i][1] == 'Female':
        target = wb.copy_worksheet(agent13Cbt)
        newsheetname = wb.get_sheet_names()[-1]
        wb.get_sheet_by_name(newsheetname).title = str(dfsorted.values[i][0]) + 'Cbt'
    else:
        target = wb.copy_worksheet(antManCbt)
        newsheetname = wb.get_sheet_names()[-1]
        wb.get_sheet_by_name(newsheetname).title = str(dfsorted.values[i][0]) + 'Cbt'

print('New sheets copied: ' + str(wb.get_sheet_names()[-len(dfsorted):]))


# In[104]:

#print('New sheets copied: ' + str(wb.get_sheet_names()[-len(characters):]))


# In[105]:

def replace_cells(character):
    for i in range(1,275):
            for j in range(1,20):
                if type(currentsheet.cell(row = i, column = j).value) == str:
                    #print(currentsheet)
                    currentsheet.cell(row = i, column = j).value = currentsheet.cell(row = i, column = j).value.replace(character, str(dfsorted.values[k][0]))
    #print('copy: ',str(character))
    print('new char copied:' ,str(dfsorted.values[k][0]),'...')

                    
                    
k = 0   
for i in range(0, len(dfsorted)):
    currentsheet = wb.get_sheet_by_name(str(wb.get_sheet_names()[-len(dfsorted) + i]))
    if dfsorted.values[i][1] == 'Female':
        replace_cells('agent13')
    else:
        replace_cells('antMan')
    k += 1
    #print(k)


# In[106]:

#crossbonesCbt =  wb.get_sheet_by_name('crossbonesCbt')
#crossbonesCbt['C7'].value


# In[107]:

#claireTempleCbt =  wb.get_sheet_by_name('claireTempleCbt1')
#claireTempleCbt['E7'].value


# In[110]:

newfilename = filename.replace('.xlsx','_updated.xlsx').replace(' ','_')
print('new file name: ',newfilename)


# In[109]:

wb.save(newfilename)


# In[ ]:



