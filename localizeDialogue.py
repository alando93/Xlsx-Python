import openpyxl
filename = input('Enter workbook filename: ')
wb = openpyxl.load_workbook(filename)

#wb.get_sheet_names()
sheetname = input('Enter dialogue sheet name: ')

#set variable to where first row and columns of character dialogues begin
r0 = int(input('Enter first row number where you want to start exporting tokens: '))
r1 = int(input('Enter last row: '))
c0 = 25


dg = wb.get_sheet_by_name(sheetname)
#[testx,testy] = [dg.cell(row = r0, column = c0).value, dg.cell(row = r0, column = 26).value]

#print(testx)
#print(testy)

locallist = []

for i in range(r0,r1):
    b = c0
    if dg.cell(row = i, column = b).value != None:
        [x,y] = [dg.cell(row = i, column = b).value, dg.cell(row = i, column = b+1).value]
        l = [x,y]
        locallist.append(l)
        #print(i, l)
    if dg.cell(row = i, column = b+2).value != None:
        [a,b] = [dg.cell(row = i, column = b+2).value, dg.cell(row = i, column = b+1+2).value]
        m = [a,b]
        locallist.append(m)
        #print(i, m)

export = openpyxl.Workbook()
export.get_sheet_names()

sheet = export.get_sheet_by_name('Sheet')
sheet.title = 'LocalData'
export.get_sheet_names()

a = 1
for i in range(0,len(locallist)-1):
    #write "LocalData"
    sheet.cell(row = a, column = 1).value = "LocalData"
    #write text token
    sheet.cell(row = a, column = 2).value = locallist[a][0]
    #write text token
    sheet.cell(row = a, column = 5).value = locallist[a][1]
    print(a,sheet.cell(row = a, column = 2).value,sheet.cell(row = a, column = 5).value)
    a += 1

export.save('localDataExport.xlsx')
